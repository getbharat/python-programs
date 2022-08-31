import openpyxl

workbook = openpyxl.load_workbook("inventory.xlsx")
supplier_inventory_sheet = workbook["Sheet1"]
products_per_supplier = {}
inventory_value_per_supplier = {}
product_out_of_stock = {}

print(supplier_inventory_sheet.max_row)

for row in range(2, supplier_inventory_sheet.max_row + 1):
    supplier_name = supplier_inventory_sheet.cell(row, 4).value
    inventory = supplier_inventory_sheet.cell(row, 2).value
    price = supplier_inventory_sheet.cell(row, 3).value
    product = supplier_inventory_sheet.cell(row, 1).value
    inventory_price = supplier_inventory_sheet.cell(row, 5)

    # product count per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        products_per_supplier[supplier_name] = 1

    # Inventory value per supplier
    if supplier_name in inventory_value_per_supplier:
        inventory_value_per_supplier[supplier_name] = inventory_value_per_supplier.get(
            supplier_name) + inventory * price
    else:
        inventory_value_per_supplier[supplier_name] = inventory * price

    # Product that has less than 10 units in inventory
    if inventory < 10:
        product_out_of_stock[int(product)] = int(inventory)

    inventory_price.value = inventory * price

print(products_per_supplier)
print(inventory_value_per_supplier)
print(product_out_of_stock)

workbook.save("inventory_solutions.xlsx")
